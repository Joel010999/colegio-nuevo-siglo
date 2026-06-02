from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
import csv
import re
from decimal import Decimal

from .models import (
    Alumno, RegistroDeuda, ConceptoDeuda, 
    PerfilUsuario, Pago, ConfiguracionSistema, RegistroAuditoria
)


# ==================== DECORADORES ====================

def admin_required(view_func):
    """Decorator que verifica si el usuario es administrador."""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('portal:login')
        if not hasattr(request.user, 'perfil') or not request.user.perfil.es_admin:
            messages.error(request, 'No tiene permisos para acceder a esta sección.')
            return redirect('portal:portal_padre')
        return view_func(request, *args, **kwargs)
    return wrapper


def check_password_change(view_func):
    """Decorator que verifica si el usuario debe cambiar su contraseña."""
    def wrapper(request, *args, **kwargs):
        if request.user.is_authenticated:
            if hasattr(request.user, 'perfil') and request.user.perfil.must_change_password:
                return redirect('portal:primer_login')
        return view_func(request, *args, **kwargs)
    return wrapper


# ==================== AUTENTICACIÓN ====================

def login_view(request):
    """Vista de login."""
    if request.user.is_authenticated:
        if hasattr(request.user, 'perfil'):
            if request.user.perfil.must_change_password:
                return redirect('portal:primer_login')
            if request.user.perfil.es_admin:
                return redirect('portal:admin_dashboard')
        return redirect('portal:portal_padre')
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            RegistroAuditoria.log(user, 'LOGIN', f'Inicio de sesión: {username}', request)
            
            # Verificar si debe cambiar contraseña
            if hasattr(user, 'perfil') and user.perfil.must_change_password:
                return redirect('portal:primer_login')
            
            # Redirigir según rol
            if hasattr(user, 'perfil') and user.perfil.es_admin:
                return redirect('portal:admin_dashboard')
            return redirect('portal:portal_padre')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
    
    return render(request, 'portal/login.html')


@login_required
def logout_view(request):
    """Vista de logout."""
    RegistroAuditoria.log(request.user, 'LOGOUT', 'Cierre de sesión', request)
    logout(request)
    messages.success(request, 'Sesión cerrada correctamente')
    return redirect('portal:login')


@login_required
def primer_login_view(request):
    """Vista para cambio de contraseña obligatorio en primer ingreso."""
    if not hasattr(request.user, 'perfil') or not request.user.perfil.must_change_password:
        return redirect('portal:portal_padre')
    
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        confirm_password = request.POST.get('confirm_password', '')
        
        # Validaciones
        email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
        if not re.match(email_regex, email):
            messages.error(request, 'Ingrese un email válido')
            return render(request, 'portal/primer_login.html')
        
        password_regex = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d).{8,20}$'
        if not re.match(password_regex, password):
            messages.error(request, 'La contraseña debe tener entre 8 y 20 caracteres, 1 mayúscula, 1 minúscula y 1 número')
            return render(request, 'portal/primer_login.html')
        
        if password != confirm_password:
            messages.error(request, 'Las contraseñas no coinciden')
            return render(request, 'portal/primer_login.html')
        
        # Actualizar usuario
        request.user.email = email
        request.user.set_password(password)
        request.user.save()
        
        # Actualizar perfil
        if hasattr(request.user, 'perfil'):
            request.user.perfil.must_change_password = False
            request.user.perfil.save()
            
            # Actualizar email de alumnos vinculados a este DNI
            if request.user.perfil.dni:
                alumnos_vinculados = Alumno.objects.filter(
                    Q(padre_dni=request.user.perfil.dni) |
                    Q(madre_dni=request.user.perfil.dni) |
                    Q(tutor_dni=request.user.perfil.dni)
                )
                
                for alumno in alumnos_vinculados:
                    modificado = False
                    if alumno.padre_dni == request.user.perfil.dni:
                        alumno.padre_email = email
                        modificado = True
                    if alumno.madre_dni == request.user.perfil.dni:
                        alumno.madre_email = email
                        modificado = True
                    if alumno.tutor_dni == request.user.perfil.dni:
                        alumno.tutor_email = email
                        modificado = True
                    
                    # Si no estaba registrado como tutor explícitamente pero el DNI del alumno
                    # coincide con el DNI del usuario logueado (caso alumnos mayores)
                    if not modificado and alumno.documento == request.user.perfil.dni:
                        alumno.email = email
                        modificado = True
                        
                    if modificado:
                        alumno.save()
        
        RegistroAuditoria.log(request.user, 'PASSWORD_CHANGED', 'Primer ingreso completado', request)
        
        # Re-autenticar
        user = authenticate(request, username=request.user.username, password=password)
        if user:
            login(request, user)
        
        messages.success(request, 'Datos actualizados correctamente')
        return redirect('portal:portal_padre')
    
    return render(request, 'portal/primer_login.html')


# ==================== PORTAL PADRES ====================

@login_required
@check_password_change
def portal_padre(request):
    """Panel principal para padres/responsables."""
    config = ConfiguracionSistema.get_config()
    
    # Obtener DNI del usuario logueado
    dni = None
    if hasattr(request.user, 'perfil') and request.user.perfil.dni:
        dni = request.user.perfil.dni
    
    alumnos_data = []
    total_general = 0
    
    if dni:
        # Buscar alumnos por DNI de responsables
        alumnos = Alumno.objects.filter(
            Q(documento=dni) |
            Q(padre_dni=dni) |
            Q(madre_dni=dni) |
            Q(tutor_dni=dni)
        ).prefetch_related('deudas', 'deudas__concepto').distinct()
        
        for alumno in alumnos:
            total_alumno = alumno.deudas.filter(estado__in=['pendiente', 'parcial'], monto__gt=0).aggregate(total=Sum('monto'))['total'] or 0
            deudas_mostrar = alumno.deudas.exclude(estado='no_corresponde').order_by('concepto__orden')
            
            alumnos_data.append({
                'alumno': alumno,
                'deudas': deudas_mostrar,
                'total': total_alumno,
            })
            total_general += total_alumno
    
    context = {
        'config': config,
        'alumnos': alumnos_data,
        'total_adeudado': total_general,
        'usuario': request.user,
    }
    
    return render(request, 'portal/portal_padre.html', context)


@login_required
@check_password_change
def enviar_comprobante(request, deuda_id):
    """Vista para enviar comprobante de pago."""
    from django.urls import reverse
    
    deuda = get_object_or_404(RegistroDeuda, id=deuda_id)
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    if request.method == 'POST':
        monto = request.POST.get('monto', '')
        comprobante = request.FILES.get('comprobante')
        
        if not comprobante:
            if is_ajax:
                return JsonResponse({'success': False, 'error': 'Debe adjuntar un comprobante'})
            messages.error(request, 'Debe adjuntar un comprobante')
            return redirect('portal:portal_padre')
        
        try:
            monto_decimal = Decimal(monto)
        except:
            monto_decimal = deuda.monto
        
        # Crear registro de pago
        pago = Pago.objects.create(
            deuda=deuda,
            monto_pagado=monto_decimal,
            comprobante=comprobante,
            usuario_responsable=request.user
        )
        
        # Actualizar estado de la deuda
        deuda.estado = 'comprobante_enviado'
        deuda.save()
        
        RegistroAuditoria.log(
            request.user, 'PAYMENT_SUBMITTED', 
            f'Pago enviado: {pago.numero_operacion} - ${monto_decimal} - {deuda.concepto.nombre}',
            request
        )
        
        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': f'Comprobante enviado correctamente. Nº Operación: {pago.numero_operacion}',
                'recibo_url': reverse('portal:ver_recibo', args=[pago.id])
            })
        
        messages.success(request, f'Comprobante enviado correctamente. Nº Operación: {pago.numero_operacion}')
        return redirect('portal:ver_recibo', pago_id=pago.id)
    
    return redirect('portal:portal_padre')


@login_required
@check_password_change
def ver_recibo(request, pago_id):
    """Vista para ver el recibo provisorio."""
    pago = get_object_or_404(Pago, id=pago_id)
    
    context = {
        'pago': pago,
        'deuda': pago.deuda,
        'alumno': pago.deuda.alumno,
    }
    
    return render(request, 'portal/recibo.html', context)


# ==================== PANEL ADMINISTRATIVO ====================

@login_required
@admin_required
def admin_dashboard(request):
    """Dashboard administrativo con estadísticas."""
    deudas_count = RegistroDeuda.objects.filter(monto__gt=0).exclude(estado__in=['no_corresponde', 'pagado', 'pago_verificado']).count()
    pagos_pendientes = Pago.objects.filter(estado='pendiente').count()
    pagos_verificados = Pago.objects.filter(estado='verificado').count()
    total_recaudado = Pago.objects.filter(estado='verificado').aggregate(
        total=Sum('monto_pagado'))['total'] or 0
    pagos_por_mes = Pago.objects.filter(estado='verificado').annotate(mes=TruncMonth('fecha_envio')).values('mes').annotate(total=Sum('monto_pagado')).order_by('-mes')
    
    # Pagos recientes pendientes de verificación
    pagos_recientes = Pago.objects.filter(estado='pendiente').order_by('-fecha_envio')[:5]
    
    context = {
        'deudas_count': deudas_count,
        'pagos_pendientes': pagos_pendientes,
        'pagos_verificados': pagos_verificados,
        'total_recaudado': total_recaudado,
        'pagos_recientes': pagos_recientes,
        'active_tab': 'dashboard',
        'pagos_por_mes': pagos_por_mes,
    }
    
    return render(request, 'portal/admin/dashboard.html', context)


@login_required
@admin_required
def admin_deudas(request):
    """Lista de deudas con filtros y estadísticas."""
    deudas = RegistroDeuda.objects.select_related('alumno', 'concepto').filter(monto__gt=0).exclude(estado__in=['no_corresponde', 'pagado']).order_by('-id')
    
    # Estadísticas para el dashboard
    deudas_count = RegistroDeuda.objects.filter(monto__gt=0).exclude(estado__in=['no_corresponde', 'pagado', 'pago_verificado']).count()
    pagos_pendientes = Pago.objects.filter(estado='pendiente').count()
    pagos_verificados = Pago.objects.filter(estado='verificado').count()
    total_recaudado = Pago.objects.filter(estado='verificado').aggregate(
        total=Sum('monto_pagado'))['total'] or 0
    pagos_por_mes = Pago.objects.filter(estado='verificado').annotate(mes=TruncMonth('fecha_envio')).values('mes').annotate(total=Sum('monto_pagado')).order_by('-mes')
    
    # Filtros
    nivel_filter = request.GET.get('nivel', '')
    curso_filter = request.GET.get('curso', '')
    division_filter = request.GET.get('division', '')
    estado_filter = request.GET.get('estado', '')
    nombre_filter = request.GET.get('nombre', '').strip()
    apellido_filter = request.GET.get('apellido', '').strip()
    dni_filter = request.GET.get('dni', '').strip()
        
    if nivel_filter:
        deudas = deudas.filter(alumno__nivel=nivel_filter)
    if curso_filter:
        deudas = deudas.filter(alumno__curso=curso_filter)
    if division_filter:
        deudas = deudas.filter(alumno__division=division_filter)
    if estado_filter:
        deudas = deudas.filter(estado=estado_filter)
    if nombre_filter:
        deudas = deudas.filter(alumno__nombres__icontains=nombre_filter)
    if apellido_filter:
        deudas = deudas.filter(alumno__apellido__icontains=apellido_filter)
    if dni_filter:
        deudas = deudas.filter(alumno__documento__icontains=dni_filter)
    
    # Obtener opciones para los filtros
    # Niveles
    niveles_map = {'I4': 'Inicial 4', 'I5': 'Inicial 5', 'P': 'Primario', 'S': 'Secundario'}
    niveles_db = set(Alumno.objects.values_list('nivel', flat=True).distinct())
    niveles_ordenados = ['I4', 'I5', 'P', 'S'] + sorted([n for n in niveles_db if n and n not in niveles_map])
    niveles = [{'val': n, 'label': niveles_map.get(n, n), 'selected': n == nivel_filter} for n in niveles_ordenados]
    
    # Cursos
    cursos_db = Alumno.objects.values_list('curso', flat=True).distinct().order_by('curso')
    cursos_map = {'1': '1ro', '2': '2do', '3': '3ro', '4': '4to', '5': '5to', '6': '6to'}
    cursos = [{'val': c, 'label': cursos_map.get(c, c), 'selected': c == curso_filter} for c in cursos_db if c]
    
    # Divisiones
    divisiones_db = Alumno.objects.values_list('division', flat=True).distinct().order_by('division')
    divisiones = [{'val': d, 'label': f"División {d}", 'selected': d == division_filter} for d in divisiones_db if d]
    
    # Estados
    estados_options = [
        ('pendiente', 'Pendiente'),
        ('comprobante_enviado', 'Comprobante Enviado'),
        ('pago_verificado', 'Pago Verificado'),
    ]
    estados = [{'val': k, 'label': v, 'selected': k == estado_filter} for k, v in estados_options]

    paginator = Paginator(deudas, 50)
    page = request.GET.get('page', 1)
    deudas_page = paginator.get_page(page)
    
    context = {
        'deudas': deudas_page,
        'niveles': niveles,
        'cursos': cursos,
        'divisiones': divisiones,
        'estados': estados,
        'nivel_filter': nivel_filter,
        'curso_filter': curso_filter,
        'division_filter': division_filter,
        'estado_filter': estado_filter,
        'nombre_filter': nombre_filter,
        'apellido_filter': apellido_filter,
        'dni_filter': dni_filter,
        'active_tab': 'deudas',
        # Estadísticas
        'deudas_count': deudas_count,
        'pagos_pendientes': pagos_pendientes,
        'pagos_verificados': pagos_verificados,
        'total_recaudado': total_recaudado,
        'pagos_por_mes': pagos_por_mes,
    }
    
    if request.GET.get('ajax') == 'true' or request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'portal/admin/partials/tabla_deudas.html', context)
        
    return render(request, 'portal/admin/deudas_final.html', context)


@login_required
@admin_required
def admin_pagos(request):
    """Lista de pagos con opción de verificar."""
    pagos = Pago.objects.select_related('deuda', 'deuda__alumno', 'usuario_responsable').all()
    
    # Estadísticas para el dashboard
    deudas_count = RegistroDeuda.objects.filter(monto__gt=0).exclude(estado__in=['no_corresponde', 'pagado', 'pago_verificado']).count()
    pagos_pendientes = Pago.objects.filter(estado='pendiente').count()
    pagos_verificados = Pago.objects.filter(estado='verificado').count()
    total_recaudado = Pago.objects.filter(estado='verificado').aggregate(
        total=Sum('monto_pagado'))['total'] or 0
    pagos_por_mes = Pago.objects.filter(estado='verificado').annotate(mes=TruncMonth('fecha_envio')).values('mes').annotate(total=Sum('monto_pagado')).order_by('-mes')
    
    # Filtros
    estado_filter = request.GET.get('estado', '')
    dni_filter = request.GET.get('dni', '').strip()
    nombre_filter = request.GET.get('nombre', '').strip()

    if estado_filter:
        pagos = pagos.filter(estado=estado_filter)
    if dni_filter:
        pagos = pagos.filter(deuda__alumno__documento__icontains=dni_filter)
    if nombre_filter:
        pagos = pagos.filter(deuda__alumno__apellido__icontains=nombre_filter)
    
    paginator = Paginator(pagos, 50)
    page = request.GET.get('page', 1)
    pagos_page = paginator.get_page(page)
    
    context = {
        'pagos': pagos_page,
        'estado_filter': estado_filter,
        'dni_filter': dni_filter,
        'nombre_filter': nombre_filter,
        'active_tab': 'pagos',
        # Estadísticas
        'deudas_count': deudas_count,
        'pagos_pendientes': pagos_pendientes,
        'pagos_verificados': pagos_verificados,
        'total_recaudado': total_recaudado,
        'pagos_por_mes': pagos_por_mes,
    }
    
    return render(request, 'portal/admin/pagos_fixed.html', context)


@login_required
@admin_required
def admin_verificar_pago(request, pago_id):
    """Verificar un pago."""
    pago = get_object_or_404(Pago, id=pago_id)
    
    if request.method == 'POST':
        accion = request.POST.get('accion', 'verificar')
        
        if accion == 'verificar':
            pago.verificar(request.user)
            RegistroAuditoria.log(
                request.user, 'PAYMENT_VERIFIED',
                f'Pago verificado: {pago.numero_operacion} - ${pago.monto_pagado}',
                request
            )
            messages.success(request, f'Pago {pago.numero_operacion} verificado correctamente')
        elif accion == 'rechazar':
            pago.estado = 'rechazado'
            pago.save()
            pago.deuda.estado = 'pendiente'
            pago.deuda.save()
            messages.warning(request, f'Pago {pago.numero_operacion} rechazado')
        
        return redirect('portal:admin_pagos')
    
    context = {
        'pago': pago,
        'active_tab': 'pagos',
    }
    
    return render(request, 'portal/admin/verificar_pago.html', context)


@login_required
@admin_required
def admin_cobrar_efectivo(request, deuda_id):
    """Cobrar una deuda en efectivo (pago rápido por ventanilla) y verificarlo asíncronamente."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
        
    deuda = get_object_or_404(RegistroDeuda, id=deuda_id)
    
    if deuda.estado == 'pendiente':
        import uuid
        numero_op = f"EFECTIVO-{timezone.now().strftime('%Y%m%d%H%M%S')}-{str(uuid.uuid4())[:4].upper()}"
        
        pago = Pago.objects.create(
            deuda=deuda,
            monto_pagado=deuda.monto,
            estado='pendiente',
            usuario_responsable=request.user,
            observaciones='Pago en Efectivo (Ventanilla)',
            numero_operacion=numero_op
        )
        pago.verificar(request.user)
        
        # Registrar en auditoría
        RegistroAuditoria.log(
            request.user, 'PAYMENT_VERIFIED',
            f'Cobro rápido en efectivo: {numero_op} - ${pago.monto_pagado} - {deuda.concepto.nombre}',
            request
        )
        
        return JsonResponse({'success': True, 'nuevo_estado': 'pago_verificado'})
        
    return JsonResponse({'success': False, 'error': 'La deuda no se encuentra en estado pendiente'}, status=400)


@login_required
@admin_required
def admin_usuarios(request):
    """Lista de alumnos por curso con estado de usuario."""
    # Obtener todos los alumnos agrupados por curso
    alumnos = Alumno.objects.all().order_by('apellido', 'nombres')
    
    # Agrupar por curso
    alumnos_por_curso = {}
    for alumno in alumnos:
        curso = alumno.curso_completo if alumno.curso_completo else 'Sin curso asignado'
        
        if curso not in alumnos_por_curso:
            alumnos_por_curso[curso] = []
        
        # Buscar si tiene usuario asociado
        perfil = PerfilUsuario.objects.select_related('usuario').filter(dni=alumno.documento).first()
        
        # Obtener email: Primero del usuario si existe, luego del tutor/responsable
        email_responsable = ''
        if perfil and perfil.usuario.email:
            email_responsable = perfil.usuario.email
        else:
            email_responsable = alumno.tutor_email or alumno.padre_email or alumno.madre_email or alumno.email or ''
        
        alumnos_por_curso[curso].append({
            'alumno': alumno,
            'nombre_completo': f"{alumno.nombres} {alumno.apellido}",
            'dni': alumno.documento,
            'email': email_responsable,
            'perfil': perfil,
            'tiene_usuario': perfil is not None,
            'must_change_password': perfil.must_change_password if perfil else True,
        })
    
    context = {
        'alumnos_por_curso': dict(sorted(alumnos_por_curso.items())),
        'active_tab': 'usuarios',
    }
    
    return render(request, 'portal/admin/usuarios.html', context)


@login_required
@admin_required
def admin_reset_password(request, usuario_id):
    """Resetear contraseña de un usuario."""
    perfil = get_object_or_404(PerfilUsuario, id=usuario_id)
    config = ConfiguracionSistema.get_config()
    
    perfil.usuario.set_password(config.password_default)
    perfil.usuario.save()
    perfil.must_change_password = True
    perfil.save()
    
    RegistroAuditoria.log(
        request.user, 'PASSWORD_RESET',
        f'Contraseña reseteada para {perfil.usuario.username}',
        request
    )
    
    messages.success(request, f'Contraseña de {perfil.usuario.username} reseteada a la genérica')
    return redirect('portal:admin_usuarios')


@login_required
@admin_required
def admin_force_password_change(request, usuario_id):
    """Forzar cambio de contraseña en el próximo login."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    perfil = get_object_or_404(PerfilUsuario, id=usuario_id)
    perfil.must_change_password = True
    perfil.save()
    
    RegistroAuditoria.log(
        request.user, 'PASSWORD_RESET',
        f'Forzado cambio de contraseña para {perfil.usuario.username}',
        request
    )
    
    return JsonResponse({'success': True})


@login_required
@admin_required
def admin_crear_alumno(request):
    """Crear un nuevo alumno y su usuario correspondiente."""
    import json
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos inválidos'})
    
    # Validar campos obligatorios
    documento = data.get('documento', '').strip()
    apellido = data.get('apellido', '').strip()
    nombres = data.get('nombres', '').strip()
    nivel = data.get('nivel', '').strip()
    curso = data.get('curso', '').strip()
    division = data.get('division', '').strip()
    tutor_email = data.get('tutor_email', '').strip()
    tutor_nombre = data.get('tutor_nombre', '').strip()
    tutor_dni = data.get('tutor_dni', '').strip()
    
    if not all([documento, apellido, nombres, nivel, curso, division, tutor_email]):
        return JsonResponse({'success': False, 'error': 'Complete todos los campos obligatorios'})
    
    # Validar DNI
    try:
        dni_alumno = int(documento)
    except ValueError:
        return JsonResponse({'success': False, 'error': 'DNI inválido'})
    
    # Verificar que el alumno no exista
    if Alumno.objects.filter(documento=dni_alumno).exists():
        return JsonResponse({'success': False, 'error': f'Ya existe un alumno con DNI {dni_alumno}'})
    
    # Verificar que el username no exista
    username = str(dni_alumno)
    if User.objects.filter(username=username).exists():
        return JsonResponse({'success': False, 'error': f'Ya existe un usuario con ese DNI'})
    
    config = ConfiguracionSistema.get_config()
    
    try:
        # Crear Alumno
        alumno = Alumno.objects.create(
            documento=dni_alumno,
            apellido=apellido,
            nombres=nombres,
            nivel=nivel,
            curso=curso,
            division=division,
            tutor_email=tutor_email,
            tutor_nombre=tutor_nombre,
            tutor_dni=int(tutor_dni) if tutor_dni else None,
        )
        
        # Crear User
        user = User.objects.create_user(
            username=username,
            password=config.password_default,
            first_name=nombres,
            last_name=apellido,
            email=tutor_email
        )
        
        # Crear PerfilUsuario
        PerfilUsuario.objects.create(
            usuario=user,
            dni=dni_alumno,
            rol='padre',
            must_change_password=True
        )
        
        # Registrar en auditoría
        RegistroAuditoria.log(
            request.user, 'USER_CREATED',
            f'Alumno y usuario creado: {apellido}, {nombres} (DNI: {dni_alumno})',
            request
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Usuario creado exitosamente',
            'dni': dni_alumno,
            'password_default': config.password_default
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al crear usuario: {str(e)}'})


@login_required
@admin_required
def admin_avisos(request):
    """Envío de avisos de deuda."""
    # Capturar filtros GET
    dni_filter = request.GET.get('dni', '').strip()
    apellido_filter = request.GET.get('apellido', '').strip()

    # Obtener morosos (deudas pendientes agrupadas por alumno)
    morosos = []
    alumnos_con_deuda = Alumno.objects.filter(
        deudas__estado='pendiente'
    ).annotate(
        total_deuda=Sum('deudas__monto', filter=Q(deudas__estado='pendiente'))
    ).distinct()

    # Aplicar filtros al queryset
    if dni_filter:
        alumnos_con_deuda = alumnos_con_deuda.filter(documento__icontains=dni_filter)
    if apellido_filter:
        alumnos_con_deuda = alumnos_con_deuda.filter(apellido__icontains=apellido_filter)
    
    for alumno in alumnos_con_deuda:
        # Buscar email del responsable
        perfil = PerfilUsuario.objects.select_related('usuario').filter(dni=alumno.documento).first()
        email = ''
        if perfil and perfil.usuario.email:
            email = perfil.usuario.email
        else:
            email = alumno.padre_email or alumno.madre_email or alumno.tutor_email or alumno.email
        
        morosos.append({
            'alumno': alumno,
            'email': email,
            'total_deuda': alumno.total_deuda or 0,
        })
    
    context = {
        'morosos': morosos,
        'config': ConfiguracionSistema.get_config(),
        'active_tab': 'avisos',
        'dni_filter': dni_filter,
        'apellido_filter': apellido_filter,
    }
    
    return render(request, 'portal/admin/avisos.html', context)


@login_required
@admin_required
def admin_enviar_avisos_masivos(request):
    """Enviar avisos masivos a todos los morosos por email.
    
    Envía el mismo mensaje genérico a todos los morosos con email.
    Lanza el envío en un hilo separado para no trabar Railway.
    """
    from .email_services import enviar_emails_masivos_async
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    asunto = request.POST.get('asunto', '').strip()
    mensaje = request.POST.get('mensaje', '').strip()
    
    if not asunto or not mensaje:
        return JsonResponse({'success': False, 'error': 'Debe completar el asunto y el mensaje'})
    
    # Obtener morosos con deuda pendiente
    alumnos_con_deuda = Alumno.objects.filter(
        deudas__estado='pendiente'
    ).distinct()
    
    destinatarios = []
    emails_sin_correo = []
    
    for alumno in alumnos_con_deuda:
        # Prioridad: email del User (actualizado en primer login) > emails del Alumno
        perfil = PerfilUsuario.objects.select_related('usuario').filter(dni=alumno.documento).first()
        email = ''
        if perfil and perfil.usuario.email:
            email = perfil.usuario.email
        else:
            email = alumno.padre_email or alumno.madre_email or alumno.tutor_email or alumno.email
        
        if email:
            destinatarios.append(email)
        else:
            emails_sin_correo.append(alumno.nombre_completo)
    
    # Deduplicar emails (un padre puede tener varios hijos)
    destinatarios = list(set(destinatarios))
    
    if not destinatarios:
        return JsonResponse({
            'success': False, 
            'error': 'No hay morosos con email registrado para enviar avisos'
        })
    
    # Construir versión HTML del mensaje con links clicables
    import re
    mensaje_html = mensaje.replace('\n', '<br>')
    # Hacer clicables las URLs del portal
    mensaje_html = re.sub(
        r'(https?://[^\s<]+)',
        r'<a href="\1" target="_blank" style="color:#1976D2;font-weight:bold;">\1</a>',
        mensaje_html
    )
    mensaje_html = f'<div style="font-family:Arial,sans-serif;font-size:15px;line-height:1.6;color:#333;">{mensaje_html}</div>'
    
    # Lanzar envío en hilo separado (no bloquea Railway)
    enviar_emails_masivos_async(
        destinatarios=destinatarios,
        asunto=asunto,
        mensaje_texto=mensaje,
        mensaje_html=mensaje_html,
        batch_size=50,
        delay=10,
        total_padres_db=len(destinatarios) + len(emails_sin_correo),
    )
    
    # Registrar en auditoría
    RegistroAuditoria.log(
        request.user, 'EMAIL_SENT',
        f'Envío masivo iniciado: {len(destinatarios)} emails en background - Asunto: {asunto[:50]}',
        request
    )
    
    return JsonResponse({
        'success': True,
        'enviados': len(destinatarios),
        'emails': destinatarios,
        'sin_correo': emails_sin_correo,
        'message': f'Envío masivo iniciado: {len(destinatarios)} emails se están enviando en segundo plano'
    })


@login_required
@admin_required
def admin_enviar_aviso_individual(request):
    """Enviar aviso individual a un moroso por email."""
    from django.core.mail import send_mail, BadHeaderError
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    email = request.POST.get('email', '').strip()
    asunto = request.POST.get('asunto', '').strip()
    mensaje = request.POST.get('mensaje', '').strip()
    
    if not email or not asunto or not mensaje:
        return JsonResponse({'success': False, 'error': 'Faltan datos requeridos (email, asunto o mensaje)'})
    
    try:
        from django.conf import settings as django_settings
        from_email = getattr(django_settings, 'DEFAULT_FROM_EMAIL', 'cobranzasns@colegionuevosiglo.edu.ar')
        
        # Enviar email
        send_mail(
            asunto,
            mensaje,
            from_email,
            [email],
            fail_silently=False,
        )
        
        # Registrar en auditoría
        RegistroAuditoria.log(
            request.user, 'EMAIL_SENT',
            f'Aviso individual enviado a {email} - Asunto: {asunto[:50]}',
            request
        )
        
        return JsonResponse({'success': True, 'message': f'Email enviado correctamente a {email}'})
        
    except BadHeaderError:
        return JsonResponse({'success': False, 'error': 'Error en el encabezado del email'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al enviar email: {str(e)}'})


@login_required
@login_required
@login_required
@admin_required
def admin_cobro(request):
    """Vista para generar deudas mensuales (Publicar Pagos)."""
    conceptos = ConceptoDeuda.objects.all().order_by('orden', 'codigo')
    
    context = {
        'active_tab': 'cobro',
        'conceptos': conceptos,
    }
    return render(request, 'portal/admin/cobro.html', context)


@login_required
@admin_required
def admin_archivos(request):
    """Vista unificada de Archivos (Importar/Exportar)."""
    # Datos para Exportar
    alumnos_count = Alumno.objects.count()
    deudas_count = RegistroDeuda.objects.filter(monto__gt=0).exclude(estado__in=['no_corresponde', 'pagado', 'pago_verificado']).count()
    total_deuda = RegistroDeuda.objects.filter(monto__gt=0).exclude(estado__in=['no_corresponde', 'pagado', 'pago_verificado']).aggregate(
        total=Sum('monto'))['total'] or 0
    
    # Contexto base
    context = {
        'alumnos_count': alumnos_count,
        'deudas_count': deudas_count,
        'total_deuda': total_deuda,
        'active_tab': 'archivos',
        'fecha_actual': timezone.now().strftime('%Y%m%d'),
        'resultados': request.session.pop('import_resultados', None) # Recuperar resultados si existen
    }
    
    return render(request, 'portal/admin/archivos.html', context)


@login_required
@admin_required
def admin_importar(request):
    """
    Procesador Unificado de Deudas - Importa desde dos archivos (Sistema Azul + Sistema Rojo).
    
    Sistema Azul: Cuotas y Matrícula (.xlsx o .csv)
    Sistema Rojo: Jornada Extendida y Materiales (.xlsx o .csv)
    
    Ambos archivos se fusionan por Documento (DNI) en un solo DataFrame consolidado
    antes de procesarse contra la base de datos.
    """
    import pandas as pd
    from io import BytesIO
    import re
    
    resultados = None
    
    if request.method == 'POST':
        # ============================================================
        # PASO 0: Recepción y validación de ambos archivos
        # ============================================================
        archivo_azul = request.FILES.get('archivo_azul')
        archivo_rojo = request.FILES.get('archivo_rojo')
        
        print("[1/4] Archivos recibidos. Leyendo DataFrames...")
        
        if not archivo_azul or not archivo_rojo:
            messages.error(request, '❌ Debe cargar ambos archivos: Sistema Azul y Sistema Rojo.')
            # Redirigir a archivos si viene de ahí
            next_url = request.POST.get('next')
            if next_url:
                return redirect(next_url)
            return redirect('portal:admin_archivos')
        
        reemplazar = request.POST.get('reemplazar') == 'on'
        
        config = ConfiguracionSistema.get_config()
        from django.contrib.auth.hashers import make_password
        hashed_default_pwd = make_password(config.password_default)
        
        added = 0
        updated = 0
        skipped = 0
        duplicados = 0
        users_created = 0
        errores = []
        
        try:
            # ============================================================
            # PASO 1: Leer ambos archivos con pandas
            # ============================================================
            def leer_archivo_a_dataframe(archivo_upload):
                """Lee un archivo Excel o CSV y devuelve un DataFrame de pandas."""
                fname = archivo_upload.name.lower()
                raw_bytes = archivo_upload.read()
                
                if fname.endswith(('.xlsx', '.xls')):
                    df = pd.read_excel(BytesIO(raw_bytes), dtype=str, engine='openpyxl')
                elif fname.endswith('.csv'):
                    # Intentar UTF-8 primero, luego latin-1
                    try:
                        content = raw_bytes.decode('utf-8-sig')
                    except UnicodeDecodeError:
                        content = raw_bytes.decode('latin-1')
                    from io import StringIO
                    # Detectar delimitador
                    sample = content.split('\n')[0] if '\n' in content else content
                    delimiter = ';' if sample.count(';') > sample.count(',') else ','
                    df = pd.read_csv(StringIO(content), delimiter=delimiter, dtype=str)
                else:
                    raise ValueError(f'Formato no soportado: {fname}. Use .xlsx, .xls o .csv')
                
                # Limpiar headers: quitar espacios
                df.columns = [str(c).strip() for c in df.columns]
                
                # Scanner dinámico: si la primera fila no tiene 'Documento',
                # buscar en las primeras 10 filas la fila de headers real
                header_cols_lower = [c.lower() for c in df.columns]
                if 'documento' not in header_cols_lower and 'apellido' not in header_cols_lower:
                    for i in range(min(10, len(df))):
                        row_vals = [str(v).strip() for v in df.iloc[i].values]
                        row_lower = [v.lower() for v in row_vals]
                        if 'documento' in row_lower and 'apellido' in row_lower:
                            # Re-leer usando esta fila como header
                            df.columns = row_vals
                            df = df.iloc[i+1:].reset_index(drop=True)
                            df.columns = [str(c).strip() for c in df.columns]
                            break
                
                return df
            
            print("[IMPORTAR] Leyendo archivo Sistema Azul...")
            df_azul = leer_archivo_a_dataframe(archivo_azul)
            print(f"[IMPORTAR] Sistema Azul: {len(df_azul)} filas, columnas: {list(df_azul.columns[:10])}...")
            
            print("[IMPORTAR] Leyendo archivo Sistema Rojo...")
            df_rojo = leer_archivo_a_dataframe(archivo_rojo)
            print(f"[IMPORTAR] Sistema Rojo: {len(df_rojo)} filas, columnas: {list(df_rojo.columns[:10])}...")
            
            # ============================================================
            # PASO 2: Mapeo de columnas del Sistema Azul
            # El archivo Azul usa "1_Cuota", "2_Cuota", etc.
            # La base de datos espera "1_Cuota Marzo", "2_Cuota Abril", etc.
            # ============================================================
            MAPEO_CUOTAS_AZUL = {
                '1_Cuota': '1_Cuota Marzo',
                '2_Cuota': '2_Cuota Abril',
                '3_Cuota': '3_Cuota Mayo',
                '4_Cuota': '4_Cuota Junio',
                '5_Cuota': '5_Cuota Julio',
                '6_Cuota': '6_Cuota Agosto',
                '7_Cuota': '7_Cuota Septiembre',
                '8_Cuota': '8_Cuota Octubre',
                '9_Cuota': '9_Cuota Noviembre',
                '10_Cuota': '10_Cuota Diciembre',
            }
            
            # Aplicar rename al DataFrame azul (solo columnas que existan)
            rename_map = {k: v for k, v in MAPEO_CUOTAS_AZUL.items() if k in df_azul.columns}
            if rename_map:
                df_azul = df_azul.rename(columns=rename_map)
                print(f"[IMPORTAR] Columnas renombradas en Azul: {rename_map}")
            
            print(f"[2/4] Sistema Azul ({len(df_azul)} filas) | Sistema Rojo ({len(df_rojo)} filas)")
            
            # ============================================================
            # PASO 3: Fusión (Merge) de ambos DataFrames por Documento
            # Usamos outer join para no perder alumnos de ningún sistema.
            # Las columnas compartidas (Apellido, Nombres, Niv, etc.) se
            # toman del Azul como prioridad (_x), con fallback al Rojo (_y).
            # ============================================================
            
            # Normalizar nombre de columna Documento
            for df_target, nombre_sistema in [(df_azul, 'Azul'), (df_rojo, 'Rojo')]:
                cols_lower = {c.lower(): c for c in df_target.columns}
                if 'documento' not in cols_lower:
                    raise ValueError(f'El archivo {nombre_sistema} no contiene la columna "Documento".')
            
            # Encontrar el nombre exacto de la columna Documento en cada DF
            doc_col_azul = next(c for c in df_azul.columns if c.lower() == 'documento')
            doc_col_rojo = next(c for c in df_rojo.columns if c.lower() == 'documento')
            
            # Estandarizar a 'Documento' en ambos
            df_azul = df_azul.rename(columns={doc_col_azul: 'Documento'})
            df_rojo = df_rojo.rename(columns={doc_col_rojo: 'Documento'})
            
            # Convertir Documento a numérico para el merge (quitar .0, espacios, etc.)
            df_azul['Documento'] = pd.to_numeric(df_azul['Documento'], errors='coerce')
            df_rojo['Documento'] = pd.to_numeric(df_rojo['Documento'], errors='coerce')
            
            # Eliminar filas sin documento válido
            df_azul = df_azul.dropna(subset=['Documento'])
            df_rojo = df_rojo.dropna(subset=['Documento'])
            df_azul['Documento'] = df_azul['Documento'].astype(int)
            df_rojo['Documento'] = df_rojo['Documento'].astype(int)
            
            # Hacer el merge (outer join para no perder alumnos)
            df_merged = pd.merge(df_azul, df_rojo, on='Documento', how='outer', suffixes=('_azul', '_rojo'))
            
            print(f"[IMPORTAR] DataFrame fusionado: {len(df_merged)} filas, {len(df_merged.columns)} columnas")
            
            # ============================================================
            # PASO 3.5: Resolver columnas duplicadas de datos del alumno
            # Prioridad: Azul > Rojo para Apellido, Nombres, Niv, Cur, Div, Familia
            # ============================================================
            COLUMNAS_ALUMNO = ['Apellido', 'Nombres', 'Niv', 'Cur', 'Div', 'Familia']
            for col_base in COLUMNAS_ALUMNO:
                col_azul = f'{col_base}_azul'
                col_rojo = f'{col_base}_rojo'
                if col_azul in df_merged.columns and col_rojo in df_merged.columns:
                    # Prioridad azul, fallback rojo
                    df_merged[col_base] = df_merged[col_azul].fillna(df_merged[col_rojo])
                    df_merged = df_merged.drop(columns=[col_azul, col_rojo])
                elif col_azul in df_merged.columns:
                    df_merged = df_merged.rename(columns={col_azul: col_base})
                elif col_rojo in df_merged.columns:
                    df_merged = df_merged.rename(columns={col_rojo: col_base})
                # Si col_base ya existe directamente (sin sufijo), no hacer nada
            
            # ============================================================
            # PASO 4: Detectar columnas de concepto (patrón dígito_nombre)
            # ============================================================
            concepto_pattern = re.compile(r'^\d+_')
            concepto_columns = [c for c in df_merged.columns if concepto_pattern.match(c)]
            
            print(f"[IMPORTAR] Columnas de concepto detectadas ({len(concepto_columns)}): {concepto_columns[:10]}...")
            
            if not concepto_columns:
                messages.error(request, '❌ No se detectaron columnas de concepto (formato N_Nombre) en los archivos fusionados.')
                next_url = request.POST.get('next')
                if next_url:
                    return redirect(next_url)
                return redirect('portal:admin_archivos')
            
            # ============================================================
            # PASO 5: Pre-registrar conceptos y procesar deudas en transacción
            # ============================================================
            from django.db import transaction
            
            with transaction.atomic():
                ConceptoDeuda.objects.all().update(orden=9999)
                
                for col_order, concepto_header in enumerate(concepto_columns):
                    c_nombre = str(concepto_header).strip()
                    if len(c_nombre) > 20:
                        c_codigo = (c_nombre[:10] + c_nombre[-10:]).upper().replace(' ', '_')
                    else:
                        c_codigo = c_nombre.upper().replace(' ', '_')
                    concepto_obj, _ = ConceptoDeuda.objects.get_or_create(
                        codigo=c_codigo,
                        defaults={'nombre': c_nombre, 'orden': col_order}
                    )
                    concepto_obj.nombre = c_nombre
                    concepto_obj.orden = col_order
                    concepto_obj.save()
                
                # ============================================================
                # PASO 6: Iterar sobre el DataFrame fusionado y procesar
                # ============================================================
                total_filas = len(df_merged)
                print("[3/4] Fusión completa. Iniciando procesamiento en Base de Datos...")
                
                for index, (row_idx, row) in enumerate(df_merged.iterrows()):
                    fila_num = row_idx + 2  # Para mensajes de error (1-indexed + header)
                    
                    if index % 50 == 0:
                        print(f"Procesando alumno {index}/{total_filas}...")
                    
                    # --- Datos del alumno ---
                    dni_val = row.get('Documento')
                    if pd.isna(dni_val):
                        errores.append(f'Fila {fila_num}: Sin documento')
                        skipped += 1
                        continue
                    
                    try:
                        dni_alumno = int(dni_val)
                    except (ValueError, TypeError):
                        errores.append(f'Fila {fila_num}: DNI inválido "{dni_val}"')
                        skipped += 1
                        continue
                    
                    # Extraer datos del alumno (case-insensitive lookup)
                    def get_col(name, default=''):
                        """Busca columna por nombre (case-insensitive)."""
                        for c in df_merged.columns:
                            if c.lower() == name.lower():
                                val = row.get(c)
                                return str(val).strip() if pd.notna(val) else default
                        return default
                    
                    apellido = get_col('Apellido')
                    nombres = get_col('Nombres')
                    nivel = get_col('Niv')
                    curso = get_col('Cur')
                    division = get_col('Div')
                    
                    # Crear/obtener alumno
                    alumno, alumno_created = Alumno.objects.get_or_create(
                        documento=dni_alumno,
                        defaults={
                            'apellido': apellido,
                            'nombres': nombres,
                            'nivel': nivel,
                            'curso': curso,
                            'division': division,
                        }
                    )
                    
                    # Actualizar datos si cambió
                    if not alumno_created:
                        updated_alumno = False
                        if apellido and alumno.apellido != apellido:
                            alumno.apellido = apellido
                            updated_alumno = True
                        if nombres and alumno.nombres != nombres:
                            alumno.nombres = nombres
                            updated_alumno = True
                        if nivel and alumno.nivel != nivel:
                            alumno.nivel = nivel
                            updated_alumno = True
                        if curso and alumno.curso != curso:
                            alumno.curso = curso
                            updated_alumno = True
                        if division and alumno.division != division:
                            alumno.division = division
                            updated_alumno = True
                        if updated_alumno:
                            alumno.save()
                    
                    # Crear usuario si no existe
                    username = str(dni_alumno)
                    if not User.objects.filter(username=username).exists():
                        user = User.objects.create(
                            username=username,
                            password=hashed_default_pwd,
                            first_name=nombres,
                            last_name=apellido
                        )
                        PerfilUsuario.objects.create(
                            usuario=user,
                            dni=dni_alumno,
                            rol='padre',
                            must_change_password=True
                        )
                        users_created += 1
                    
                    # --- Procesar cada columna de concepto (BLINDAJE DE DEUDAS) ---
                    for concepto_header in concepto_columns:
                        monto_val = row.get(concepto_header)
                        
                        # Filtrar vacíos, nulos y ceros
                        if pd.isna(monto_val) or str(monto_val).strip() == '' or monto_val == 0 or str(monto_val).strip() == '0':
                            continue
                        
                        # Código único del concepto
                        c_nombre = str(concepto_header).strip()
                        if len(c_nombre) > 20:
                            concepto_codigo = (c_nombre[:10] + c_nombre[-10:]).upper().replace(' ', '_')
                        else:
                            concepto_codigo = c_nombre.upper().replace(' ', '_')
                        
                        concepto, _ = ConceptoDeuda.objects.get_or_create(
                            codigo=concepto_codigo,
                            defaults={'nombre': c_nombre, 'orden': 0}
                        )
                        
                        # --- Validar valores especiales de texto ---
                        val_str = str(monto_val).lower().strip()
                        
                        if 'pagad' in val_str:
                            deuda_existente = RegistroDeuda.objects.filter(
                                alumno=alumno, concepto=concepto
                            ).first()
                            
                            if deuda_existente:
                                # REGLA DE ORO: proteger estados verificados
                                if deuda_existente.estado in ('pagado', 'pago_verificado', 'comprobante_enviado'):
                                    skipped += 1
                                    continue
                                if reemplazar and deuda_existente.estado == 'pendiente':
                                    deuda_existente.monto = 0
                                    deuda_existente.estado = 'pagado'
                                    deuda_existente.save()
                                    updated += 1
                                else:
                                    duplicados += 1
                            else:
                                RegistroDeuda.objects.create(
                                    alumno=alumno, concepto=concepto,
                                    monto=0, periodo='', estado='pagado'
                                )
                                added += 1
                            continue
                        
                        if 'no corresponde' in val_str or 'nocorresponde' in val_str:
                            deuda_existente = RegistroDeuda.objects.filter(
                                alumno=alumno, concepto=concepto
                            ).first()
                            
                            if deuda_existente:
                                if deuda_existente.estado in ('pagado', 'pago_verificado', 'comprobante_enviado'):
                                    skipped += 1
                                    continue
                                if reemplazar and deuda_existente.estado == 'pendiente':
                                    deuda_existente.monto = 0
                                    deuda_existente.estado = 'no_corresponde'
                                    deuda_existente.save()
                                    updated += 1
                                else:
                                    duplicados += 1
                            else:
                                RegistroDeuda.objects.create(
                                    alumno=alumno, concepto=concepto,
                                    monto=0, periodo='', estado='no_corresponde'
                                )
                                added += 1
                            continue
                        
                        # --- Valor numérico ---
                        try:
                            monto = Decimal(str(monto_val).replace(',', '.').strip())
                            if monto <= 0:
                                continue
                        except:
                            continue
                        
                        # Verificar duplicado (mismo alumno + concepto)
                        # order_by('-id') es CRÍTICO para ver siempre el estado de la última cuota generada
                        deuda_existente = RegistroDeuda.objects.filter(
                            alumno=alumno, concepto=concepto
                        ).order_by('-id').first()
                        
                        CONCEPTOS_RECURRENTES = [
                            '25_INGLES PETS', '26_INGLES FCES', '27_HORA COMPLEMENTARIA',
                            '29_JORNADA EXTENDIDA', '31_CUOTA ADMISION', '31_CUOTA FEBRERO',
                            '33_JORNADA EXT 3 DIAS', '33_JORNADA EXTEND 3 DIA', 
                            '35_INGLES KET', '36_INGLES MATERIAL', '37_OLIMPIADAS'
                        ]
                        
                        if deuda_existente:
                            # Si la última deuda generada está pagada/verificada...
                            if deuda_existente.estado in ('pagado', 'pago_verificado', 'comprobante_enviado'):
                                
                                es_recurrente = any(c_rec in concepto_header for c_rec in CONCEPTOS_RECURRENTES)
                                
                                if es_recurrente:
                                    # Opción 1: Crear una NUEVA deuda (cajoncito nuevo) para este concepto recurrente
                                    RegistroDeuda.objects.create(
                                        alumno=alumno, concepto=concepto,
                                        monto=monto, periodo='', estado='pendiente'
                                    )
                                    added += 1
                                else:
                                    # REGLA DE ORO NORMAL: Proteger y omitir
                                    skipped += 1
                                    continue
                                    
                            # Si la última deuda sigue pendiente, actualizar para absorber recargos
                            elif deuda_existente.estado == 'pendiente':
                                if deuda_existente.monto != monto:
                                    deuda_existente.monto = monto
                                    deuda_existente.save()
                                    updated += 1
                                else:
                                    duplicados += 1
                            else:
                                duplicados += 1
                        else:
                            # Si no existe ninguna deuda previa, crearla normalmente
                            RegistroDeuda.objects.create(
                                alumno=alumno, concepto=concepto,
                                monto=monto, periodo='', estado='pendiente'
                            )
                            added += 1
            
            print("[4/4] Importación exitosa. Redirigiendo...")
            
            # ============================================================
            # PASO 7: Resultados y auditoría
            # ============================================================
            resultados = {
                'added': added,
                'updated': updated,
                'skipped': skipped,
                'duplicados': duplicados,
                'users_created': users_created,
                'errores': errores[:20],
                'total_errores': len(errores),
            }
            
            RegistroAuditoria.log(
                request.user, 'IMPORT',
                f'Importación Unificada (Azul+Rojo): {added} nuevas, {updated} actualizadas, '
                f'{duplicados} duplicados, {skipped} omitidas, {users_created} usuarios',
                request
            )
            
            if added > 0 or updated > 0:
                messages.success(
                    request,
                    f'✅ Importación completada: {added} deudas nuevas, {updated} actualizadas, '
                    f'{users_created} usuarios creados'
                )
            if duplicados > 0:
                messages.warning(request, f'⚠️ {duplicados} registros duplicados no fueron importados (ya existen)')
            if skipped > 0:
                messages.warning(request, f'⚠️ {skipped} registros omitidos (protegidos o con errores)')
                
        except Exception as e:
            import traceback
            print(f"[IMPORTAR] ERROR: {traceback.format_exc()}")
            messages.error(request, f'Error al procesar los archivos: {str(e)}')
    
    # Guardar resultados en sesión y redirigir a archivos
    if resultados:
        request.session['import_resultados'] = resultados
    
    next_url = request.POST.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('portal:admin_archivos')


def procesar_fila_estandar(row_idx, row, config, reemplazar, hashed_default_pwd):
    """Procesa una fila en formato estándar (una deuda por fila)."""
    result = {'status': 'error', 'error': '', 'user_created': False}
    
    # Mapear columnas flexibles
    alumno_nombre = (
        row.get('alumno') or row.get('nombre_alumno') or 
        row.get('estudiante') or row.get('nombre') or ''
    )
    
    dni_str = (
        row.get('dni_alumno') or row.get('dni') or 
        row.get('documento') or row.get('doc') or ''
    ).replace('.', '').replace('-', '')
    
    curso = (
        row.get('curso') or row.get('division') or 
        row.get('curso/division') or row.get('grado') or ''
    )
    
    concepto_nombre = (
        row.get('concepto') or row.get('descripcion') or 
        row.get('detalle') or 'Cuota'
    )
    
    monto_str = (
        row.get('monto') or row.get('importe') or 
        row.get('monto_adeudado') or row.get('deuda') or '0'
    )
    
    periodo = (
        row.get('periodo') or row.get('mes') or 
        row.get('mes_periodo') or row.get('fecha') or ''
    )
    
    # Validar DNI
    if not dni_str:
        result['error'] = f'Fila {row_idx}: Sin DNI - {alumno_nombre}'
        return result
    
    try:
        dni_alumno = int(dni_str)
    except ValueError:
        result['error'] = f'Fila {row_idx}: DNI inválido "{dni_str}"'
        return result
    
    # Validar monto
    try:
        monto = Decimal(monto_str.replace(',', '.').replace('$', '').strip())
    except:
        result['error'] = f'Fila {row_idx}: Monto inválido "{monto_str}"'
        return result
    
    # Parsear nombre
    if ',' in alumno_nombre:
        apellido = alumno_nombre.split(',')[0].strip()
        nombres = alumno_nombre.split(',')[1].strip() if len(alumno_nombre.split(',')) > 1 else ''
    else:
        partes = alumno_nombre.split()
        apellido = partes[0] if partes else ''
        nombres = ' '.join(partes[1:]) if len(partes) > 1 else ''
    
    # Obtener o crear alumno
    alumno, alumno_created = Alumno.objects.get_or_create(
        documento=dni_alumno,
        defaults={'apellido': apellido, 'nombres': nombres, 'curso': curso}
    )
    
    if not alumno_created and curso and alumno.curso != curso:
        alumno.curso = curso
        alumno.save()
    
    # Crear usuario si no existe
    username = str(dni_alumno)
    if not User.objects.filter(username=username).exists():
        user = User.objects.create(
            username=username,
            password=hashed_default_pwd,
            first_name=nombres,
            last_name=apellido
        )
        PerfilUsuario.objects.create(
            usuario=user,
            dni=dni_alumno,
            rol='padre',
            must_change_password=True
        )
        result['user_created'] = True
    
    # Obtener o crear concepto
    concepto, _ = ConceptoDeuda.objects.get_or_create(
        codigo=concepto_nombre[:20].upper().replace(' ', '_'),
        defaults={'nombre': concepto_nombre}
    )
    
    # Verificar duplicado
    deuda_existente = RegistroDeuda.objects.filter(
        alumno=alumno,
        concepto=concepto,
        periodo=periodo
    ).first()
    
    if deuda_existente:
        # Proteger pagos verificados/comprobantes enviados
        if deuda_existente.estado in ('pago_verificado', 'comprobante_enviado'):
            result['status'] = 'error'
            result['error'] = f'Fila {row_idx}: Pago ya verificado, no se modifica'
            return result
        if reemplazar and deuda_existente.estado == 'pendiente':
            deuda_existente.monto = monto
            deuda_existente.save()
            result['status'] = 'updated'
        else:
            result['status'] = 'duplicado'
    else:
        RegistroDeuda.objects.create(
            alumno=alumno,
            concepto=concepto,
            monto=monto,
            periodo=periodo,
            estado='pendiente'
        )
        result['status'] = 'added'
    
    return result


@login_required
@admin_required
def admin_exportar(request):
    """Exportar deudas a Excel o CSV con formato pivoteado."""
    from io import BytesIO
    from decimal import Decimal
    from django.conf import settings
    import os
    from openpyxl.utils import get_column_letter

    if request.method == 'POST':
        formato = request.POST.get('formato', 'csv')
        
        # 1. Obtener conceptos activos (los del ultimo Excel importado, orden < 9000)
        conceptos = ConceptoDeuda.objects.filter(orden__lt=9000).order_by('orden')
        
        # 2. Obtener todos los alumnos
        alumnos = Alumno.objects.all().order_by('apellido', 'nombres')
        
        # 3. Obtener TODAS las deudas (Snapshot en tiempo real)
        all_deudas = RegistroDeuda.objects.select_related('alumno', 'concepto').all()
        
        # 4. Mapear deudas en memoria para acceso rápido
        # Estructura: deudas_map[alumno_pk] = { concepto_id: valor }
        # Valor puede ser el monto (si es pendiente) o 'pagado' (si está verificado)
        deudas_map = {}
        for d in all_deudas:
            if d.alumno_id not in deudas_map:
                deudas_map[d.alumno_id] = {}
            
            # Prioridad: Si ya hay un estado 'pagado' para este concepto, se mantiene.
            # Si no, se guarda el actual.
            current_val = deudas_map[d.alumno_id].get(d.concepto_id)
            
            # Si ya está marcado como texto especial, no sobrescribir excepto si hay pago real?
            if current_val in ['pagado', 'no corresponde']:
                continue

            if d.estado in ['pago_verificado', 'pagado']:
                deudas_map[d.alumno_id][d.concepto_id] = 'pagado'
            elif d.estado == 'no_corresponde':
                deudas_map[d.alumno_id][d.concepto_id] = 'no corresponde'
            else:
                # Si es pendiente/otro, sumamos el monto
                if isinstance(current_val, (int, float, Decimal)):
                    deudas_map[d.alumno_id][d.concepto_id] += d.monto
                else:
                    deudas_map[d.alumno_id][d.concepto_id] = d.monto
        
        # Headers del archivo
        headers = ['Familia', 'Documento', 'Apellido', 'Nombres', 'Niv', 'Cur', 'Div', 'Saldo_Moroso']
        # Agregar headers dinámicos de conceptos
        # Si c.nombre ya tiene patrón "dígito_", usarlo directo (viene del import colegio)
        header_cols = []
        for c in conceptos:
            if re.match(r'^\d+_', c.nombre):
                header_cols.append(c.nombre)
            else:
                header_cols.append(f"{c.codigo}_{c.nombre}")
        headers.extend(header_cols)
        
        # Construir filas
        rows = []
        for alumno in alumnos:
            datos_deuda = deudas_map.get(alumno.pk, {})
            
            # Fila base
            row = [
                alumno.familia,
                alumno.documento,
                alumno.apellido,
                alumno.nombres,
                alumno.nivel,
                alumno.curso,
                alumno.division,
                0, # Placeholder para Saldo_Moroso (Index 7, Col H)
            ]
            
            # Rellenar columnas de conceptos
            for concepto in conceptos:
                val = datos_deuda.get(concepto.id, 0.0)
                # Excel prefiere números o string 'pagado'
                if val == 'pagado':
                    row.append('pagado')
                elif val == 'no corresponde':
                    row.append('no corresponde')
                else:
                    row.append(float(val))
            
            rows.append(row)
        
        # Crear directorio de exportaciones si no existe
        export_dir = os.path.join(settings.MEDIA_ROOT, 'exportaciones')
        os.makedirs(export_dir, exist_ok=True)
        
        fecha_str = timezone.now().strftime('%Y%m%d_%H%M%S')
        
        if formato == 'excel':
            # Generar Excel
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Deudas"
            
            # Escribir headers con estilo
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="800020", end_color="800020", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Escribir datos
            last_col_idx = len(headers)
            last_col_letter = get_column_letter(last_col_idx)
            start_sum_col = 9 # Columna I (Conceptos empiezan acá)
            start_sum_letter = get_column_letter(start_sum_col)
            
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    # Si es la columna H (Saldo Moroso, index 8), poner fórmula
                    if col_idx == 8:
                        # Formula: =SUM(I{row}:<Last>{row})
                        formula = f"=SUM({start_sum_letter}{row_idx}:{last_col_letter}{row_idx})"
                        ws.cell(row=row_idx, column=col_idx, value=formula)
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Ajustar anchos de columna
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 30)
            
            filename = f'deudas_{fecha_str}.xlsx'
            filepath = os.path.join(export_dir, filename)
            wb.save(filepath)
            
        else:
            # Generar CSV
            filename = f'deudas_{fecha_str}.csv'
            filepath = os.path.join(export_dir, filename)
            
            with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for row_data in rows:
                    # Calcular saldo para CSV manual
                    saldo = 0
                    csv_row = list(row_data) # Copia
                    
                    # Sumar conceptos (Indices 8 en adelante)
                    for val in csv_row[8:]:
                        if isinstance(val, (int, float)):
                            saldo += val
                    
                    csv_row[7] = saldo # Actualizar Col H
                    writer.writerow(csv_row)
        
        RegistroAuditoria.log(
            request.user, 'EXPORT',
            f'Exportación de deudas ({len(rows)} alumnos) en formato {formato.upper()}',
            request
        )
        
        # Redirigir a la descarga del archivo
        from django.http import FileResponse
        file_handle = open(filepath, 'rb')
        response = FileResponse(file_handle, as_attachment=True, filename=filename)
        return response
    
    # GET - Mostrar página
    alumnos_count = Alumno.objects.count()
    deudas_count = RegistroDeuda.objects.filter(estado='pendiente').count()
    total_deuda = RegistroDeuda.objects.filter(estado='pendiente').aggregate(
        total=Sum('monto'))['total'] or 0
    
    return redirect('portal:admin_archivos')


@login_required
@admin_required
def admin_config(request):
    """Configuración del sistema."""
    config = ConfiguracionSistema.get_config()
    
    if request.method == 'POST':
        config.alias_transferencia = request.POST.get('alias', config.alias_transferencia)
        config.cbu = request.POST.get('cbu', config.cbu)
        config.password_default = request.POST.get('password_default', config.password_default)
        config.save()
        
        RegistroAuditoria.log(request.user, 'CONFIG_UPDATE', 'Configuración actualizada', request)
        messages.success(request, 'Configuración guardada correctamente')
    
    # Obtener usuarios admin
    admins = PerfilUsuario.objects.filter(rol='admin').select_related('usuario')
    
    context = {
        'config': config,
        'admins': admins,
        'active_tab': 'config',
    }
    
    return render(request, 'portal/admin/config.html', context)


@login_required
@admin_required
def admin_auditoria(request):
    """Log de auditoría."""
    from datetime import datetime
    
    fecha = request.GET.get('fecha', '')
    usuario = request.GET.get('usuario', '')

    registros = RegistroAuditoria.objects.select_related('usuario').all()
    
    if fecha:
        try:
            dt = datetime.strptime(fecha, '%Y-%m-%d')
            registros = registros.filter(timestamp__date=dt.date())
        except ValueError:
            pass
            
    if usuario:
        registros = registros.filter(usuario__username__icontains=usuario)

    registros = registros[:200]
    
    context = {
        'registros': registros,
        'active_tab': 'auditoria',
        'filtro_fecha': fecha,
        'filtro_usuario': usuario,
    }
    
    return render(request, 'portal/admin/auditoria.html', context)


@user_passes_test(lambda u: u.is_superuser)
def reset_database_nuclear(request):
    try:
        # Borrado en orden para respetar claves foráneas
        Pago.objects.all().delete()
        RegistroDeuda.objects.all().delete()
        ConceptoDeuda.objects.all().delete()
        Alumno.objects.all().delete()
        
        # Borrar usuarios normales (padres), preservando superusuarios y staff
        User.objects.filter(is_superuser=False, is_staff=False).delete()
        
        return HttpResponse('''
            <div style="font-family: sans-serif; text-align: center; margin-top: 50px;">
                <h1 style="color: #2ecc71;">✅ Base de datos aniquilada con éxito</h1>
                <p>Todo el sistema está limpio. Las columnas viejas y los usuarios de prueba fueron eliminados.</p>
                <p>El administrador principal sigue intacto.</p>
                <a href="/admin-panel/" style="padding: 10px 20px; background: #3498db; color: white; text-decoration: none; border-radius: 5px;">Volver al Panel Admin</a>
            </div>
        ''')
    except Exception as e:
        return HttpResponse(f"<h1 style='color: red;'>❌ Error al limpiar:</h1><p>{str(e)}</p>")

# ==================== VISTAS PÚBLICAS (consulta sin login) ====================

def consulta_publica(request):
    """
    Vista pública para consultar deudas sin login.
    Permite buscar por DNI del alumno o responsable.
    """
    context = {
        'alumnos': [],
        'total_adeudado': 0,
        'dni_buscado': '',
        'mensaje_error': '',
        'busqueda_realizada': False,
    }
    
    dni = request.GET.get('dni', '').strip()
    
    if dni:
        context['dni_buscado'] = dni
        context['busqueda_realizada'] = True
        
        try:
            dni_int = int(dni)
            
            alumnos = Alumno.objects.filter(
                Q(documento=dni_int) |
                Q(padre_dni=dni_int) |
                Q(madre_dni=dni_int) |
                Q(tutor_dni=dni_int)
            ).prefetch_related('deudas', 'deudas__concepto').distinct()
            
            if alumnos.exists():
                alumnos_data = []
                total_general = 0
                
                for alumno in alumnos:
                    total_alumno = alumno.deudas.filter(estado__in=['pendiente', 'parcial'], monto__gt=0).aggregate(total=Sum('monto'))['total'] or 0
                    
                    alumnos_data.append({
                        'alumno': alumno,
                        'deudas': alumno.deudas.all(),
                        'total': total_alumno,
                    })
                    total_general += total_alumno
                
                context['alumnos'] = alumnos_data
                context['total_adeudado'] = total_general
            else:
                context['mensaje_error'] = f'No se encontraron registros para el DNI {dni}'
                
        except ValueError:
            context['mensaje_error'] = 'Por favor ingrese un DNI válido (solo números)'
    
    return render(request, 'portal/consulta_publica.html', context)


# ==================== TEST EMAIL ====================

@login_required
@admin_required
def test_email_batch(request):
    """
    Vista de prueba para verificar el sistema de envío masivo.
    Envía 5 emails de prueba usando el servicio de batching.
    Solo accesible para administradores.
    """
    from .email_services import enviar_emails_masivos
    from django.conf import settings as django_settings

    # 5 emails de prueba (cambiar por direcciones propias)
    emails_prueba = [
        'joeljjs100@gmail.com',
        'clientmagnetweb@gmail.com',
    ]

    resultado = enviar_emails_masivos(
        destinatarios=emails_prueba,
        asunto='[TEST] Prueba de envío masivo — Colegio Nuevo Siglo',
        mensaje_texto=(
            'Este es un email de prueba del sistema de envío masivo.\n'
            'Si recibís este mensaje, el sistema funciona correctamente.\n\n'
            'Colegio Nuevo Siglo'
        ),
        mensaje_html=(
            '<h2>Prueba de Envío Masivo</h2>'
            '<p>Este es un email de prueba del sistema de envío masivo.</p>'
            '<p>Si recibís este mensaje, el sistema funciona correctamente.</p>'
            '<br><p><strong>Colegio Nuevo Siglo</strong></p>'
        ),
        batch_size=1,
        delay=30,
    )

    backend_actual = django_settings.EMAIL_BACKEND
    test_mode = getattr(django_settings, 'EMAIL_TEST_MODE', True)

    return JsonResponse({
        'success': True,
        'modo': 'CONSOLA (test)' if test_mode else 'SMTP REAL',
        'backend': backend_actual,
        'resultado': resultado,
    })
